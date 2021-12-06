# -*- coding: utf-8 -*-
"""
Created on Sun Dec  5 15:09:05 2021

@author: asus
"""

from openpyxl import Workbook
from openpyxl import load_workbook
import urllib.request
import pandas as pd
from bs4 import BeautifulSoup
from string import ascii_uppercase


def write_to(this_sheet, array_to_write, file_pos):
    file_pos += 1
    print(file_pos)
    for j in range(len(array_to_write)):
        this_sheet[ascii_uppercase[j] + str(file_pos)].value = array_to_write[j]
    return file_pos

def get_medals(medal) -> tuple:
    gold = 0
    silver = 0
    bronze = 0
    if medal != None:
        if medal == "GOLD":
            gold = 1
        if medal == "SILVER":
            silver = 1
        if medal == "BRONZE":
            bronze = 1
    return (gold, silver, bronze)


workbook = Workbook()


sheet = workbook.active


#athWb = load_workbook(filename="new_datasets/olympic_results.xlsx")
athWb = load_workbook(filename="new_datasets/olympic_results.xlsx")

print("loaded")
ath_sheet = athWb.active
names = []

new_deletion = []


file_pos = 0

total = len(ath_sheet["A"])

for i in range(len(ath_sheet["A"])):
    if i == 0:
        a = ["Sport", "Discipline", "GameType", "Gold Medal", "Silver Medal", "Bronze Medal", "Olympic Committee", "Athlete Name", "Athlete URL"]
        write_to(sheet, a, file_pos)    
        continue
    if ath_sheet["J"][i].value == None:
        aths = ath_sheet["F"][i].value
        if aths != None:
            athletes = []
            x = aths.split(",")
            y = []
            for t in x:
                t = t.replace("(", "");
                t = t.replace("[", "")
                t = t.replace(")", "")
                t = t.replace("]", "")
                t = t.replace("\'", "")
                t.strip()
                y.append(t)
            x.clear()
            for j in y:
                j = j.strip()
                x.append(j)
            for k in range(0, len(x), 2):
                athletes.append((x[k], x[k+1]))
            medals = get_medals(ath_sheet["E"][i].value)
            for athlete in athletes:
                a = ath_sheet["A"][i].value, ath_sheet["B"][i].value, ath_sheet["D"][i].value, medals[0], medals[1], medals[2], ath_sheet["I"][i].value, athlete[0], athlete[1]
                print(file_pos)
                file_pos = write_to(sheet, a, file_pos)
                print(i, total)
        else:
            a = ath_sheet["A"][i].value, ath_sheet["B"][i].value, ath_sheet["D"][i].value, medals[0], medals[1], medals[2], ath_sheet["I"][i].value, "", ""
            file_pos = write_to(sheet, a, file_pos)
        continue
    medals = get_medals(ath_sheet["E"][i].value)
        
    a = ath_sheet["A"][i].value, ath_sheet["B"][i].value, ath_sheet["D"][i].value, medals[0], medals[1], medals[2], ath_sheet["I"][i].value, ath_sheet["K"][i].value, ath_sheet["J"][i].value
    print(i,total)
    file_pos = write_to(sheet, a, file_pos)


    

workbook.save(filename = "final_datasets/events.xlsx")