# -*- coding: utf-8 -*-
"""
Created on Mon Dec  6 00:02:56 2021

@author: asus
"""

from openpyxl import Workbook
from openpyxl import load_workbook
import urllib.request
import pandas as pd
from bs4 import BeautifulSoup
from string import ascii_uppercase





workbook = Workbook()


sheet = workbook.active


athWb = load_workbook(filename="final_datasets/athletes_olympic_committee.xlsx")
#athWb = load_workbook(filename="new_datasets/DummyExcel.xlsx")

print("loaded")
ath_sheet = athWb.active

file_pos = 0

"""    ATHLETE MERGER
total = (len(ath_sheet["A"])+2)//3

for i in range((len(ath_sheet["A"])+2)//3):
    if i == 0:
        sheet["A" + "1"].value = ath_sheet["A"][0].value
        sheet["B" + "1"].value = ath_sheet["B"][0].value
        sheet["C" + "1"].value = ath_sheet["C"][0].value
        sheet["D" + "1"].value = ath_sheet["D"][0].value
        sheet["E" + "1"].value = ath_sheet["E"][0].value
        sheet["F" + "1"].value = ath_sheet["F"][0].value
        sheet["G" + "1"].value = ath_sheet["G"][0].value
        sheet["H" + "1"].value = ath_sheet["H"][0].value
    else:
        #print(ath_sheet["B"][3*i-2].value)
        sheet["A" + str(i+1)].value = ath_sheet["A"][3*i-2].value
        sheet["B" + str(i+1)].value = ath_sheet["B"][3*i-2].value
        sheet["C" + str(i+1)].value = ath_sheet["C"][3*i-2].value
        sheet["D" + str(i+1)].value = ath_sheet["D"][3*i-2].value
        sheet["E" + str(i+1)].value = ath_sheet["E"][3*i-2].value
        sheet["F" + str(i+1)].value = ath_sheet["F"][3*i-1].value
        sheet["G" + str(i+1)].value = ath_sheet["G"][3*i-1].value
        sheet["H" + str(i+1)].value = ath_sheet["H"][3*i-1].value
        sheet["I" + str(i+1)].value = ath_sheet["I"][3*i-1].value
        sheet["J" + str(i+1)].value = ath_sheet["J"][3*i].value
        sheet["K" + str(i+1)].value = ath_sheet["K"][3*i].value
        sheet["L" + str(i+1)].value = ath_sheet["L"][3*i].value
    print(i, total)
"""

total = (len(ath_sheet["A"])+1)//2

for i in range((len(ath_sheet["A"])+1)//2):
    if i == 0:
        sheet["A" + "1"].value = ath_sheet["A"][0].value
        sheet["B" + "1"].value = ath_sheet["B"][0].value
        sheet["C" + "1"].value = ath_sheet["C"][0].value
    else:
        #print(ath_sheet["B"][3*i-2].value)
        sheet["A" + str(i+1)].value = ath_sheet["A"][2*i].value
        sheet["B" + str(i+1)].value = ath_sheet["B"][2*i-1].value
        sheet["C" + str(i+1)].value = ath_sheet["C"][2*i-1].value
    print(i, total)


    

workbook.save(filename = "final_datasets/athletes_olympic_committee_ord.xlsx")