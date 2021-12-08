# -*- coding: utf-8 -*-
"""
Created on Wed Dec  8 02:33:12 2021

@author: asus
"""

from openpyxl import Workbook
from openpyxl import load_workbook
import urllib.request
import pandas as pd
from bs4 import BeautifulSoup
from string import ascii_uppercase





workbook = Workbook()


sheet = workbook.active


athWb = load_workbook(filename="final_datasets/athletes_olympic_committee.xlsx")
#athWb = load_workbook(filename="new_datasets/DummyExcel.xlsx")

print("loaded")
ath_sheet = athWb.active


total = (len(ath_sheet["A"])+1)//2

for i in range((len(ath_sheet["A"])+1)//2):
    if i == 0:
        sheet["A" + "1"].value = ath_sheet["A"][0].value
        sheet["B" + "1"].value = ath_sheet["B"][0].value
        sheet["C" + "1"].value = ath_sheet["C"][0].value
    else:
        #print(ath_sheet["B"][3*i-2].value)
        sheet["A" + str(i+1)].value = ath_sheet["A"][2*i].value
        sheet["B" + str(i+1)].value = ath_sheet["B"][2*i-1].value
        sheet["C" + str(i+1)].value = ath_sheet["C"][2*i-1].value
    print(i, total)


    

workbook.save(filename = "final_datasets/athletes_olympic_committee_ord.xlsx")