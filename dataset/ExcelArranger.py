"""
Created on Sat Dec  4 22:17:54 2021

@author: asus
"""

from openpyxl import Workbook
from openpyxl import load_workbook
import urllib.request
import pandas as pd
from bs4 import BeautifulSoup

workbook = Workbook()


sheet = workbook.active
#sheet["A1"] = "hello"
#sheet["B1"] = "world!"


#workbook.save(filename = "FinalDatasets/hello_world.xlsx")



athWb = load_workbook(filename="new_datasets/olympic_athletes(Short).xlsx")

print("loaded")
ath_sheet = athWb.active
names = []

new_deletion = []

print(len(ath_sheet["A"]))
for i in range(len(ath_sheet["A"])):
    belongs = False
    if i == 0 or ath_sheet["C"][i].value == "Tokyo 2020":
        continue
    ath = ath_sheet["A"][i]
    if ath.value:
        req = urllib.request.Request(ath.value, headers={"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.75 Safari/537.36","X-Requested-With": "XMLHttpRequest"})
        page = urllib.request.urlopen(req)
        bs = BeautifulSoup(page,"html.parser")
        birth_year = -1
        games = bs.main.find_all("h2", {"class" : "games__text text-body-xl"})
        for game in games:
            if game.text == "Tokyo 2020":
                belongs = True
    print(i,": TRUE") if belongs else print(i, ": FALSE")
    if not belongs:
        new_deletion.append(ath_sheet["A"][i].value)
print("end")

