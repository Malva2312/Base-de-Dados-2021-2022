# -*- coding: utf-8 -*-
"""
Created on Sun Dec  5 20:01:22 2021

@author: asus
"""
from openpyxl import Workbook
from openpyxl import load_workbook
import urllib.request
import pandas as pd
from bs4 import BeautifulSoup
from string import ascii_uppercase


workbook = Workbook()

sheet = workbook.active

athWb = load_workbook(filename="final_datasets/events.xlsx")
#athWb = load_workbook(filename="new_datasets/DummyExcel.xlsx")

print("loaded")
ath_sheet = athWb.active

ath_dict = {}

file_pos = 1

total = len(ath_sheet["A"])

for i in range(len(ath_sheet["A"])):
    ath_url = ath_sheet["I"][i].value
    if ath_url != None:
        if ath_url in ath_dict.keys():
            sheet["H" + str(ath_dict[ath_url])].value = sheet["H" +  str(ath_dict[ath_url])].value + ath_sheet["D"][i].value
            sheet["I" + str(ath_dict[ath_url])].value = sheet["I" +  str(ath_dict[ath_url])].value + ath_sheet["E"][i].value
            sheet["J" +  str(ath_dict[ath_url])].value = sheet["J" +  str(ath_dict[ath_url])].value + ath_sheet["F"][i].value
            print(i, "duplicate")
            continue
        ath_dict[ath_url] = file_pos
        sheet["G" + str(file_pos)].value = file_pos
        sheet["A" + str(file_pos)].value = ath_url
        sheet["H" + str(file_pos)].value = ath_sheet["D"][i].value
        sheet["I" + str(file_pos)].value = ath_sheet["E"][i].value
        sheet["J" + str(file_pos)].value = ath_sheet["F"][i].value
        print(file_pos, ath_url)
        file_pos += 1
    print(i, total)


    
workbook.save(filename = "final_datasets/hello_world.xlsx")
#workbook.save(filename = "final_datasets/athletes_id_url.xlsx")